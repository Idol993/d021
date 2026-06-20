import csv
import io
import uuid
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Table,
    TableStyle,
    Spacer,
    Image as RLImage,
)
from sqlalchemy import and_

from .audit import audit_logger
from .config import config
from .models import (
    ApprovalRecord,
    ApprovalStatus,
    GrayReleaseRecord,
    PreCheckResult,
    ReleaseRequest,
    ReleaseStatus,
    RollbackDrill,
    RollbackRecord,
    WeeklyReport,
    ensure_utc,
    get_session,
    get_utc_now,
)
from .notifier import NotificationSeverity, notifier


plt.rcParams["font.sans-serif"] = ["Microsoft YaHei", "SimHei", "Arial Unicode MS"]
plt.rcParams["axes.unicode_minus"] = False


class DrillResultDTO:
    def __init__(
        self,
        success: bool,
        duration_seconds: float,
        details: Dict[str, Any],
        issues: List[str],
        improvements: List[str],
    ):
        self.success = success
        self.duration_seconds = duration_seconds
        self.details = details
        self.issues = issues
        self.improvements = improvements


class ReportsEngine:
    def __init__(self):
        self.reports_cfg = config.get("reports", {})
        self.drill_cfg = config.get("rollback_drill", {})
        self.audit_cfg = config.get("audit", {})

    def generate_weekly_report(
        self,
        week_start: Optional[datetime] = None,
        week_end: Optional[datetime] = None,
        force: bool = False,
    ) -> Optional[WeeklyReport]:
        session = get_session()
        try:
            if week_start is None or week_end is None:
                now = get_utc_now()
                monday = now - timedelta(days=now.weekday())
                week_end = monday - timedelta(days=1)
                week_start = week_end - timedelta(days=6)
                week_start = week_start.replace(hour=0, minute=0, second=0, microsecond=0)
                week_end = week_end.replace(hour=23, minute=59, second=59, microsecond=0)

            existing = (
                session.query(WeeklyReport)
                .filter(
                    WeeklyReport.week_start == week_start,
                    WeeklyReport.week_end == week_end,
                )
                .first()
            )
            if existing and not force:
                return existing

            releases = (
                session.query(ReleaseRequest)
                .filter(
                    ReleaseRequest.submitted_at >= week_start,
                    ReleaseRequest.submitted_at <= week_end,
                )
                .all()
            )

            rollbacks = (
                session.query(RollbackRecord)
                .filter(
                    RollbackRecord.started_at >= week_start,
                    RollbackRecord.started_at <= week_end,
                    RollbackRecord.is_drill == False,
                )
                .all()
            )

            metrics = self._calculate_weekly_metrics(releases, rollbacks)
            release_details = [self._serialize_release(r) for r in releases]
            rollback_details = [self._serialize_rollback(r) for r in rollbacks]

            report = WeeklyReport(
                week_start=week_start,
                week_end=week_end,
                metrics=metrics,
                release_details=release_details,
                rollback_details=rollback_details,
            )
            session.add(report)
            session.flush()

            output_dir = Path(self.reports_cfg.get("weekly", {}).get("output_dir", "./reports/weekly"))
            output_dir.mkdir(parents=True, exist_ok=True)
            base_name = f"weekly_report_{week_start.strftime('%Y%m%d')}_{week_end.strftime('%Y%m%d')}"

            file_paths: Dict[str, str] = {}
            formats = self.reports_cfg.get("weekly", {}).get("formats", ["pdf", "xlsx"])

            if "xlsx" in formats:
                xlsx_path = output_dir / f"{base_name}.xlsx"
                self._generate_weekly_xlsx(str(xlsx_path), report)
                file_paths["xlsx"] = str(xlsx_path)

            if "pdf" in formats:
                pdf_path = output_dir / f"{base_name}.pdf"
                chart_paths: Dict[str, str] = {}
                if self.reports_cfg.get("weekly", {}).get("include_trend_charts"):
                    chart_paths = self._generate_trend_charts(output_dir, base_name, release_details, rollback_details)
                self._generate_weekly_pdf(str(pdf_path), report, chart_paths)
                file_paths["pdf"] = str(pdf_path)

            report.file_paths = file_paths
            session.commit()
            session.refresh(report)

            audit_logger.log(
                actor="system",
                action="生成周报",
                category="report",
                details={
                    "report_id": report.id,
                    "week_start": week_start.isoformat(),
                    "week_end": week_end.isoformat(),
                    "files": file_paths,
                },
            )

            notifier.send(
                title=f"📊 MES发布周报 {week_start.strftime('%m/%d')}-{week_end.strftime('%m/%d')}",
                content=(
                    f"**发布概览**\n"
                    f"- 发布申请: {metrics['total_releases']} 件\n"
                    f"- 发布成功率: {metrics['success_rate_pct']:.1f}%\n"
                    f"- 回滚次数: {metrics['rollback_count_real']} 次\n"
                    f"- 平均审批时长: {metrics['avg_approval_hours']:.1f} 小时\n"
                    f"- 熔断触发: {metrics['circuit_breaker_triggers']} 次\n"
                    f"\n文件已生成，请查收附件。"
                ),
                severity=NotificationSeverity.INFO,
                extra={"周报ID": report.id},
                attachments=list(file_paths.values()),
            )

            return report
        finally:
            session.close()

    def _calculate_weekly_metrics(
        self,
        releases: List[ReleaseRequest],
        rollbacks: List[RollbackRecord],
    ) -> Dict[str, Any]:
        total = len(releases)
        if total == 0:
            return {
                "total_releases": 0,
                "success_count": 0,
                "success_rate_pct": 100.0,
                "rollback_count_real": 0,
                "rollback_count_drill": 0,
                "precheck_pass_rate_pct": 100.0,
                "approval_reject_rate_pct": 0.0,
                "avg_approval_hours": 0.0,
                "avg_gray_duration_min": 0.0,
                "circuit_breaker_triggers": 0,
                "avg_rollback_duration_seconds": 0.0,
                "by_type": {"regular": {"total": 0, "success": 0}, "hotfix": {"total": 0, "success": 0}},
            }

        success_count = sum(
            1 for r in releases if r.status in (ReleaseStatus.FULL_RELEASED, ReleaseStatus.COMPLETED, ReleaseStatus.GRAY_COMPLETED)
        )
        precheck_passed = sum(1 for r in releases if r.status != ReleaseStatus.PRE_CHECK_FAILED)
        approval_rejected = sum(1 for r in releases if r.status == ReleaseStatus.APPROVAL_REJECTED)

        avg_approval_hours = 0.0
        approval_count = 0
        for r in releases:
            r_sub = ensure_utc(r.submitted_at)
            for a in r.approvals:
                if a.submitted_at:
                    a_sub = ensure_utc(a.submitted_at)
                    if r_sub and a_sub:
                        delta = (a_sub - r_sub).total_seconds() / 3600.0
                        if delta > 0:
                            avg_approval_hours += delta
                            approval_count += 1
        if approval_count:
            avg_approval_hours /= approval_count

        avg_gray_min = 0.0
        gray_count = 0
        cb_triggers = 0
        for r in releases:
            for g in r.gray_records:
                if g.completed_at and g.started_at:
                    g_start = ensure_utc(g.started_at)
                    g_end = ensure_utc(g.completed_at)
                    if g_start and g_end:
                        avg_gray_min += (g_end - g_start).total_seconds() / 60.0
                        gray_count += 1
                if g.circuit_breaker_triggered:
                    cb_triggers += 1
        if gray_count:
            avg_gray_min /= gray_count

        real_rb = [rb for rb in rollbacks if not rb.is_drill]
        drill_rb = [rb for rb in rollbacks if rb.is_drill]
        avg_rb_sec = 0.0
        if real_rb:
            durations = [rb.duration_seconds or 0 for rb in real_rb if rb.duration_seconds is not None]
            if durations:
                avg_rb_sec = sum(durations) / len(durations)

        by_type: Dict[str, Dict[str, int]] = {"regular": {"total": 0, "success": 0}, "hotfix": {"total": 0, "success": 0}}
        for r in releases:
            key = r.release_type.value
            if key not in by_type:
                by_type[key] = {"total": 0, "success": 0}
            by_type[key]["total"] += 1
            if r.status in (ReleaseStatus.FULL_RELEASED, ReleaseStatus.COMPLETED, ReleaseStatus.GRAY_COMPLETED):
                by_type[key]["success"] += 1

        return {
            "total_releases": total,
            "success_count": success_count,
            "success_rate_pct": round(success_count / total * 100, 1),
            "rollback_count_real": len(real_rb),
            "rollback_count_drill": len(drill_rb),
            "precheck_pass_rate_pct": round(precheck_passed / total * 100, 1),
            "approval_reject_rate_pct": round(approval_rejected / total * 100, 1),
            "avg_approval_hours": round(avg_approval_hours, 1),
            "avg_gray_duration_min": round(avg_gray_min, 1),
            "circuit_breaker_triggers": cb_triggers,
            "avg_rollback_duration_seconds": round(avg_rb_sec, 1),
            "by_type": by_type,
        }

    def _serialize_release(self, r: ReleaseRequest) -> Dict[str, Any]:
        return {
            "id": r.id,
            "version": r.version,
            "release_type": r.release_type.value,
            "title": r.title,
            "submitter": r.submitter,
            "submitted_at": r.submitted_at.isoformat() if r.submitted_at else None,
            "status": r.status.value,
            "pre_checks": [
                {"item": p.check_item.value, "passed": p.passed, "score": p.score}
                for p in r.pre_checks
            ],
            "approvals": [
                {
                    "role": a.approver_role,
                    "dept": a.approver_department,
                    "status": a.status.value,
                    "submitted_at": a.submitted_at.isoformat() if a.submitted_at else None,
                }
                for a in r.approvals
            ],
            "gray_phases": [
                {
                    "order": g.phase_order,
                    "name": g.phase_name,
                    "status": g.status,
                    "triggered_cb": g.circuit_breaker_triggered,
                }
                for g in r.gray_records
            ],
            "rollbacks": [
                {
                    "id": rb.id,
                    "type": rb.rollback_type,
                    "success": rb.success,
                    "duration_s": rb.duration_seconds,
                }
                for rb in r.rollback_records
            ],
        }

    def _serialize_rollback(self, rb: RollbackRecord) -> Dict[str, Any]:
        return {
            "id": rb.id,
            "release_id": rb.release_id,
            "from_version": rb.from_version,
            "to_version": rb.to_version,
            "trigger_source": rb.trigger_source,
            "reason": rb.reason,
            "affected_lines": rb.affected_production_lines,
            "affected_batches": rb.affected_batch_ranges,
            "started_at": rb.started_at.isoformat() if rb.started_at else None,
            "completed_at": rb.completed_at.isoformat() if rb.completed_at else None,
            "duration_seconds": rb.duration_seconds,
            "success": rb.success,
            "health_check_passed": rb.health_check_passed,
            "is_drill": rb.is_drill,
        }

    def _generate_weekly_xlsx(self, path: str, report: WeeklyReport):
        wb = Workbook()
        ws = wb.active
        ws.title = "概览"

        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        ws["A1"] = f"MES系统发布周报 - {report.week_start.strftime('%Y-%m-%d')} 至 {report.week_end.strftime('%Y-%m-%d')}"
        ws["A1"].font = Font(bold=True, size=16, color="1F4E78")
        ws.merge_cells("A1:F1")
        ws["A1"].alignment = center_align

        metrics = report.metrics or {}
        overview_data = [
            ["指标", "数值"],
            ["发布申请总数", metrics.get("total_releases", 0)],
            ["发布成功率(%)", metrics.get("success_rate_pct", 0)],
            ["回滚次数(实际)", metrics.get("rollback_count_real", 0)],
            ["回滚次数(演练)", metrics.get("rollback_count_drill", 0)],
            ["前置校验通过率(%)", metrics.get("precheck_pass_rate_pct", 0)],
            ["审批驳回率(%)", metrics.get("approval_reject_rate_pct", 0)],
            ["平均审批时长(小时)", metrics.get("avg_approval_hours", 0)],
            ["平均灰度时长(分钟)", metrics.get("avg_gray_duration_min", 0)],
            ["熔断触发次数", metrics.get("circuit_breaker_triggers", 0)],
            ["平均回滚耗时(秒)", metrics.get("avg_rollback_duration_seconds", 0)],
        ]
        row = 3
        for d in overview_data:
            for col, val in enumerate(d, 1):
                cell = ws.cell(row=row, column=col, value=val)
                if row == 3:
                    cell.font = header_font
                    cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border
            row += 1

        by_type = metrics.get("by_type", {})
        row += 1
        ws.cell(row=row, column=1, value="按发布类型").font = Font(bold=True, size=13)
        ws.merge_cells(f"A{row}:F{row}")
        row += 1
        type_headers = ["类型", "总数", "成功数", "成功率(%)"]
        for col, val in enumerate(type_headers, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        row += 1
        for key, vals in by_type.items():
            total = vals.get("total", 0)
            succ = vals.get("success", 0)
            rate = round(succ / total * 100, 1) if total else 0
            for col, val in enumerate([key, total, succ, rate], 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.alignment = center_align
                cell.border = thin_border
            row += 1

        ws2 = wb.create_sheet("发布明细")
        release_headers = ["版本", "类型", "标题", "提交人", "提交时间", "状态", "前置校验", "审批进度", "回滚次数"]
        for col, val in enumerate(release_headers, 1):
            cell = ws2.cell(row=1, column=col, value=val)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        row = 2
        for rd in report.release_details:
            precheck_str = "/".join(
                ["✓" if p["passed"] else "✗" for p in rd["pre_checks"]]
            ) or "-"
            approval_total = len(rd["approvals"])
            approval_done = sum(1 for a in rd["approvals"] if a["status"] in ("approved", "post_signed"))
            approval_str = f"{approval_done}/{approval_total}" if approval_total else "-"
            vals = [
                rd["version"],
                rd["release_type"],
                rd["title"],
                rd["submitter"],
                rd["submitted_at"],
                rd["status"],
                precheck_str,
                approval_str,
                len(rd["rollbacks"]),
            ]
            for col, val in enumerate(vals, 1):
                cell = ws2.cell(row=row, column=col, value=val)
                cell.alignment = center_align
                cell.border = thin_border
            row += 1

        ws3 = wb.create_sheet("回滚明细")
        rb_headers = ["ID", "版本(原→回)", "触发源", "影响产线数", "影响批次数", "耗时(秒)", "状态", "健康检查", "演练"]
        for col, val in enumerate(rb_headers, 1):
            cell = ws3.cell(row=1, column=col, value=val)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        row = 2
        for rd in report.rollback_details:
            vals = [
                rd["id"][:8],
                f"{rd['from_version']}→{rd['to_version']}",
                rd["trigger_source"],
                len(rd["affected_lines"]),
                len(rd["affected_batches"]),
                rd["duration_seconds"],
                "成功" if rd["success"] else "失败",
                "通过" if rd["health_check_passed"] else "未通过",
                "是" if rd["is_drill"] else "否",
            ]
            for col, val in enumerate(vals, 1):
                cell = ws3.cell(row=row, column=col, value=val)
                cell.alignment = center_align
                cell.border = thin_border
            row += 1

        for sheet in (ws, ws2, ws3):
            for col_cells in sheet.columns:
                max_len = 10
                for c in col_cells:
                    try:
                        if c.value:
                            max_len = max(max_len, min(50, len(str(c.value)) + 2))
                    except Exception:
                        pass
                sheet.column_dimensions[c.column_letter].width = max_len
            sheet.freeze_panes = "A2"

        wb.save(path)

    def _generate_trend_charts(
        self,
        output_dir: Path,
        base_name: str,
        release_details: List[Dict[str, Any]],
        rollback_details: List[Dict[str, Any]],
    ) -> Dict[str, str]:
        paths: Dict[str, str] = {}
        days: Dict[str, Dict[str, int]] = {}

        for rd in release_details:
            if not rd["submitted_at"]:
                continue
            d = rd["submitted_at"][:10]
            days.setdefault(d, {"releases": 0, "rollbacks": 0})
            days[d]["releases"] += 1

        for rd in rollback_details:
            if not rd["started_at"]:
                continue
            d = rd["started_at"][:10]
            days.setdefault(d, {"releases": 0, "rollbacks": 0})
            days[d]["rollbacks"] += 1

        sorted_days = sorted(days.keys())
        if not sorted_days:
            return paths

        releases_daily = [days[d]["releases"] for d in sorted_days]
        rollbacks_daily = [days[d]["rollbacks"] for d in sorted_days]

        fig, ax1 = plt.subplots(figsize=(10, 5))
        x = list(range(len(sorted_days)))
        ax1.bar(x, releases_daily, color="#2E86AB", label="发布数", alpha=0.8, width=0.4)
        ax2 = ax1.twinx()
        ax2.plot(x, rollbacks_daily, color="#F24236", marker="o", linewidth=2, label="回滚数")
        ax1.set_xticks(x)
        ax1.set_xticklabels([d[5:] for d in sorted_days], rotation=30)
        ax1.set_ylabel("发布数", color="#2E86AB")
        ax2.set_ylabel("回滚数", color="#F24236")
        ax1.set_title("每日发布与回滚趋势")
        lines1, labels1 = ax1.get_legend_handles_labels()
        lines2, labels2 = ax2.get_legend_handles_labels()
        ax1.legend(lines1 + lines2, labels1 + labels2, loc="upper left")
        plt.tight_layout()
        trend_path = output_dir / f"{base_name}_trend.png"
        plt.savefig(str(trend_path), dpi=120, bbox_inches="tight")
        plt.close(fig)
        paths["trend"] = str(trend_path)

        if release_details:
            statuses: Dict[str, int] = {}
            for rd in release_details:
                s = rd["status"]
                statuses[s] = statuses.get(s, 0) + 1
            fig, ax = plt.subplots(figsize=(8, 6))
            labels = list(statuses.keys())
            values = list(statuses.values())
            colors_list = ["#55A630", "#2E86AB", "#F24236", "#FFBA08", "#707070"]
            while len(colors_list) < len(labels):
                colors_list.append(colors_list[len(colors_list) % 5])
            ax.pie(values, labels=labels, autopct="%1.1f%%", colors=colors_list[:len(labels)], startangle=90)
            ax.set_title("发布状态分布")
            plt.tight_layout()
            pie_path = output_dir / f"{base_name}_pie.png"
            plt.savefig(str(pie_path), dpi=120, bbox_inches="tight")
            plt.close(fig)
            paths["pie"] = str(pie_path)

        return paths

    def _generate_weekly_pdf(self, path: str, report: WeeklyReport, chart_paths: Dict[str, str]):
        doc = SimpleDocTemplate(path, pagesize=A4, leftMargin=2 * cm, rightMargin=2 * cm, topMargin=2 * cm, bottomMargin=2 * cm)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "WeeklyTitle", parent=styles["Title"],
            fontSize=20, textColor=colors.HexColor("#1F4E78"), spaceAfter=18,
        )
        h2_style = ParagraphStyle(
            "H2", parent=styles["Heading2"],
            fontSize=14, textColor=colors.HexColor("#305496"), spaceBefore=14, spaceAfter=8,
        )
        body_style = styles["BodyText"]

        story = []
        story.append(Paragraph(
            f"MES系统发布周报<br/>{report.week_start.strftime('%Y-%m-%d')} 至 {report.week_end.strftime('%Y-%m-%d')}",
            title_style,
        ))

        story.append(Paragraph("一、核心指标概览", h2_style))
        m = report.metrics or {}
        overview_data = [
            ["指标", "数值", "指标", "数值"],
            ["发布总数", str(m.get("total_releases", 0)), "发布成功率", f"{m.get('success_rate_pct', 0)}%"],
            ["回滚(实际)", str(m.get("rollback_count_real", 0)), "回滚(演练)", str(m.get("rollback_count_drill", 0))],
            ["校验通过率", f"{m.get('precheck_pass_rate_pct', 0)}%", "审批驳回率", f"{m.get('approval_reject_rate_pct', 0)}%"],
            ["平均审批(小时)", f"{m.get('avg_approval_hours', 0)}", "平均灰度(分钟)", f"{m.get('avg_gray_duration_min', 0)}"],
            ["熔断触发次数", str(m.get("circuit_breaker_triggers", 0)), "平均回滚(秒)", f"{m.get('avg_rollback_duration_seconds', 0)}"],
        ]
        tbl = Table(overview_data, hAlign="LEFT", colWidths=[4.5 * cm, 3.5 * cm, 4.5 * cm, 3.5 * cm])
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#305496")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ]))
        story.append(tbl)

        if chart_paths.get("trend"):
            story.append(Paragraph("二、趋势分析", h2_style))
            story.append(RLImage(chart_paths["trend"], width=16 * cm, height=8 * cm))
            story.append(Spacer(1, 0.4 * cm))

        if chart_paths.get("pie"):
            story.append(RLImage(chart_paths["pie"], width=12 * cm, height=9 * cm))
            story.append(Spacer(1, 0.4 * cm))

        story.append(Paragraph("三、发布明细", h2_style))
        if report.release_details:
            header = ["版本", "类型", "标题", "提交人", "状态"]
            rows = [header]
            for rd in report.release_details:
                rows.append([
                    rd["version"],
                    rd["release_type"],
                    (rd["title"] or "")[:30],
                    rd["submitter"],
                    rd["status"],
                ])
            tbl2 = Table(rows, hAlign="LEFT", colWidths=[3 * cm, 2 * cm, 5.5 * cm, 2.5 * cm, 3.5 * cm])
            tbl2.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#305496")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
            ]))
            story.append(tbl2)
        else:
            story.append(Paragraph("本周无发布记录。", body_style))

        if report.rollback_details:
            story.append(Paragraph("四、回滚明细", h2_style))
            header = ["ID", "版本变化", "触发源", "耗时(s)", "状态"]
            rows = [header]
            for rd in report.rollback_details:
                rows.append([
                    rd["id"][:8],
                    f"{rd['from_version']}→{rd['to_version']}",
                    rd["trigger_source"],
                    str(rd["duration_seconds"]),
                    "成功" if rd["success"] else "失败",
                ])
            tbl3 = Table(rows, hAlign="LEFT", colWidths=[2.5 * cm, 3.5 * cm, 3.5 * cm, 2.5 * cm, 2.5 * cm])
            tbl3.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#C5504B")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
            ]))
            story.append(tbl3)

        doc.build(story)

    def schedule_rollback_drill(
        self,
        scheduled_at: datetime,
        title: Optional[str] = None,
        scenario: Optional[str] = None,
    ) -> RollbackDrill:
        session = get_session()
        try:
            if not title:
                title = f"定期回滚演练-{scheduled_at.strftime('%Y-%m')}"
            if not scenario:
                scenario = "模拟无菌产线灰度阶段触发熔断，自动回滚至上一稳定版本，验证数据一致性及业务恢复。"

            participants = [
                {"role": p["role"], "confirmed": False}
                for p in self.drill_cfg.get("participants", [])
            ]

            drill = RollbackDrill(
                title=title,
                scheduled_at=scheduled_at,
                status="scheduled",
                participants=participants,
                scenario_description=scenario,
            )
            session.add(drill)
            session.commit()
            session.refresh(drill)

            audit_logger.log(
                actor="system",
                action="创建回滚演练计划",
                category="drill",
                details={
                    "drill_id": drill.id,
                    "title": drill.title,
                    "scheduled_at": scheduled_at.isoformat(),
                },
            )

            notify_hours = self.drill_cfg.get("notify_before_hours", 72)
            notifier.send(
                title="📝 MES回滚演练排期通知",
                content=(
                    f"**演练标题**: {title}\n"
                    f"**计划时间**: {scheduled_at.isoformat()}\n"
                    f"**参与角色**: {', '.join(p['role'] for p in participants)}\n"
                    f"**演练场景**: {scenario}\n\n"
                    f"系统将在演练前{notify_hours}小时再次提醒。请相关人员确认参与。"
                ),
                severity=NotificationSeverity.INFO,
                extra={"演练ID": drill.id},
            )

            return drill
        finally:
            session.close()

    def execute_rollback_drill(self, drill_id: str) -> Tuple[bool, str]:
        session = get_session()
        try:
            drill = session.query(RollbackDrill).filter_by(id=drill_id).first()
            if not drill:
                return False, "演练计划不存在"

            drill.status = "running"
            drill.started_at = get_utc_now()
            session.commit()

            audit_logger.log(
                actor="system",
                action="开始执行回滚演练",
                category="drill",
                details={"drill_id": drill.id},
            )

            notifier.send(
                title="🎬 MES回滚演练开始执行",
                content=(
                    f"**演练**: {drill.title}\n"
                    f"**场景**: {drill.scenario_description or '-'}\n"
                    f"请各参与人员关注执行情况。"
                ),
                severity=NotificationSeverity.INFO,
                extra={"演练ID": drill.id},
            )

            result = self._simulate_drill_execution(drill)
            drill.completed_at = get_utc_now()
            if drill.started_at and drill.completed_at:
                started_utc = ensure_utc(drill.started_at)
                completed_utc = ensure_utc(drill.completed_at)
                drill.duration_seconds = (completed_utc - started_utc).total_seconds()
            drill.success = result.success
            drill.results = result.details
            drill.issues_found = result.issues
            drill.improvement_actions = result.improvements
            drill.status = "success" if result.success else "failed"
            session.commit()
            session.refresh(drill)

            audit_logger.log(
                actor="system",
                action="回滚演练完成",
                category="drill",
                details={
                    "drill_id": drill.id,
                    "success": drill.success,
                    "duration": drill.duration_seconds,
                    "issues": drill.issues_found,
                    "improvements": drill.improvement_actions,
                },
            )

            severity = NotificationSeverity.SUCCESS if drill.success else NotificationSeverity.WARNING
            notifier.send(
                title=f"{'✅' if drill.success else '⚠️'} MES回滚演练{'成功' if drill.success else '存在问题'}",
                content=(
                    f"**演练**: {drill.title}\n"
                    f"**耗时**: {round(drill.duration_seconds or 0, 1)} 秒\n"
                    f"**结果**: {'通过' if drill.success else '未通过'}\n\n"
                    f"**发现问题** ({len(drill.issues_found or [])}项):\n"
                    + "\n".join(f"- {i}" for i in (drill.issues_found or []))
                    + "\n\n**改进建议**:\n"
                    + "\n".join(f"- {i}" for i in (drill.improvement_actions or []))
                ),
                severity=severity,
                extra={"演练ID": drill.id},
            )
            return drill.success if drill.success is not None else False, drill.status
        finally:
            session.close()

    def _simulate_drill_execution(self, drill: RollbackDrill) -> DrillResultDTO:
        import time as _time
        _time.sleep(0.3)
        duration = 18.5 + (hash(drill.id) % 150) / 10
        criteria = self.drill_cfg.get("success_criteria", {})
        max_dur = criteria.get("max_rollback_duration_min", 30) * 60

        issues: List[str] = []
        improvements: List[str] = []
        success = True

        if duration > max_dur:
            issues.append(f"演练耗时{round(duration, 1)}秒超过阈值{max_dur}秒")
            improvements.append("优化回滚脚本并行度，缩短产线服务重启等待时间")
            success = False

        data_ok = random_like_deterministic(hash(drill.id) + 1, 0.08)
        if not data_ok:
            issues.append("演练过程中发现批记录字段数据不一致（EBR模板版本 v2.3.1）")
            improvements.append("在回滚流程中增加 EBR 模板版本一致性强校验")
            success = False

        svc_ok = random_like_deterministic(hash(drill.id) + 2, 0.05)
        if not svc_ok:
            issues.append("回滚后无菌产线设备通讯服务健康检查一度降级")
            improvements.append("回滚健康检查脚本增加 OPC-UA 握手重试机制")
            success = False

        if not issues:
            issues.append("未发现严重问题")

        details = {
            "scenario_steps": [
                {"step": "隔离R&D-01产线流量", "status": "ok", "duration_s": 2.1},
                {"step": "触发熔断阈值（模拟生产异常率5.2%）", "status": "ok", "duration_s": 1.3},
                {"step": "调用回滚API恢复至 v3.8.2", "status": "ok", "duration_s": 8.5},
                {"step": "执行DB快照一致性比对", "status": "ok" if data_ok else "warning", "duration_s": 3.2},
                {"step": "服务健康检查与设备握手验证", "status": "ok" if svc_ok else "degraded", "duration_s": 3.4},
            ],
            "duration_seconds": duration,
            "max_duration_threshold_s": max_dur,
            "data_consistency": data_ok,
            "service_availability": criteria.get("service_availability_after", 100) if svc_ok else 95,
        }

        return DrillResultDTO(
            success=success,
            duration_seconds=duration,
            details=details,
            issues=issues,
            improvements=improvements,
        )

    def query_history(
        self,
        start_time: Optional[datetime] = None,
        end_time: Optional[datetime] = None,
        production_line: Optional[str] = None,
        version: Optional[str] = None,
        include_rollbacks: bool = True,
        export_format: Optional[str] = None,
        export_dir: Optional[str] = None,
    ) -> Dict[str, Any]:
        session = get_session()
        try:
            q = session.query(ReleaseRequest)
            conditions = []
            if start_time:
                conditions.append(ReleaseRequest.submitted_at >= start_time)
            if end_time:
                conditions.append(ReleaseRequest.submitted_at <= end_time)
            if version:
                conditions.append(ReleaseRequest.version.like(f"%{version}%"))
            if conditions:
                q = q.filter(and_(*conditions))

            releases = q.order_by(ReleaseRequest.submitted_at.desc()).all()
            if production_line:
                releases = [
                    r for r in releases
                    if any(production_line in (g.production_lines or []) for g in r.gray_records)
                ]

            release_list = [self._serialize_release(r) for r in releases]
            rollback_list = []
            if include_rollbacks:
                rb_q = session.query(RollbackRecord)
                rb_conditions = []
                if start_time:
                    rb_conditions.append(RollbackRecord.started_at >= start_time)
                if end_time:
                    rb_conditions.append(RollbackRecord.started_at <= end_time)
                if rb_conditions:
                    rb_q = rb_q.filter(and_(*rb_conditions))
                if production_line:
                    pass
                rollbacks = rb_q.order_by(RollbackRecord.started_at.desc()).all()
                if production_line:
                    rollbacks = [
                        rb for rb in rollbacks
                        if production_line in (rb.affected_production_lines or [])
                    ]
                rollback_list = [self._serialize_rollback(rb) for rb in rollbacks]

            result = {
                "query": {
                    "start_time": start_time.isoformat() if start_time else None,
                    "end_time": end_time.isoformat() if end_time else None,
                    "production_line": production_line,
                    "version": version,
                },
                "counts": {"releases": len(release_list), "rollbacks": len(rollback_list)},
                "releases": release_list,
                "rollbacks": rollback_list,
            }

            if export_format and (export_format in ("csv", "xlsx")):
                out_dir = Path(export_dir or "./exports")
                out_dir.mkdir(parents=True, exist_ok=True)
                ts = get_utc_now().strftime("%Y%m%d%H%M%S")
                base_name = f"mes_release_export_{ts}"
                if export_format == "csv":
                    csv_path = out_dir / f"{base_name}_releases.csv"
                    self._export_releases_csv(release_list, str(csv_path))
                    csv_rb = out_dir / f"{base_name}_rollbacks.csv"
                    self._export_rollbacks_csv(rollback_list, str(csv_rb))
                    result["exported_files"] = [str(csv_path), str(csv_rb)]
                elif export_format == "xlsx":
                    xlsx_path = out_dir / f"{base_name}.xlsx"
                    self._export_history_xlsx(release_list, rollback_list, str(xlsx_path))
                    result["exported_files"] = [str(xlsx_path)]

            return result
        finally:
            session.close()

    def _export_releases_csv(self, items: List[Dict[str, Any]], path: str):
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(["ID", "版本", "类型", "标题", "提交人", "提交时间", "状态"])
            for r in items:
                writer.writerow([
                    r["id"], r["version"], r["release_type"], r["title"],
                    r["submitter"], r["submitted_at"], r["status"],
                ])

    def _export_rollbacks_csv(self, items: List[Dict[str, Any]], path: str):
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(["ID", "发布ID", "原版本", "回滚版本", "触发源", "开始时间", "耗时秒", "成功", "演练"])
            for r in items:
                writer.writerow([
                    r["id"], r["release_id"], r["from_version"], r["to_version"],
                    r["trigger_source"], r["started_at"], r["duration_seconds"],
                    r["success"], r["is_drill"],
                ])

    def _export_history_xlsx(
        self,
        releases: List[Dict[str, Any]],
        rollbacks: List[Dict[str, Any]],
        path: str,
    ):
        wb = Workbook()
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        ws1 = wb.active
        ws1.title = "发布记录"
        h1 = ["ID", "版本", "类型", "标题", "提交人", "提交时间", "状态"]
        for col, v in enumerate(h1, 1):
            cell = ws1.cell(row=1, column=col, value=v)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        for i, r in enumerate(releases, 2):
            for col, v in enumerate([r["id"], r["version"], r["release_type"], r["title"], r["submitter"], r["submitted_at"], r["status"]], 1):
                c = ws1.cell(row=i, column=col, value=v)
                c.border = thin_border

        ws2 = wb.create_sheet("回滚记录")
        h2 = ["ID", "发布ID", "原→回版本", "触发源", "影响产线", "影响批次", "开始时间", "耗时秒", "成功"]
        for col, v in enumerate(h2, 1):
            cell = ws2.cell(row=1, column=col, value=v)
            cell.font = header_font
            cell.fill = PatternFill(start_color="C5504B", end_color="C5504B", fill_type="solid")
            cell.border = thin_border
        for i, r in enumerate(rollbacks, 2):
            vals = [
                r["id"], r["release_id"], f"{r['from_version']}→{r['to_version']}",
                r["trigger_source"],
                ",".join(r["affected_lines"] or []),
                ",".join(r["affected_batches"] or []),
                r["started_at"], r["duration_seconds"], r["success"],
            ]
            for col, v in enumerate(vals, 1):
                c = ws2.cell(row=i, column=col, value=v)
                c.border = thin_border
        for sheet in (ws1, ws2):
            sheet.freeze_panes = "A2"
            for col_cells in sheet.columns:
                mx = 12
                for c in col_cells:
                    if c.value:
                        mx = max(mx, min(48, len(str(c.value)) + 2))
                sheet.column_dimensions[c.column_letter].width = mx
        wb.save(path)


def random_like_deterministic(seed: int, bad_ratio: float) -> bool:
    bucket = (seed & 0xFFFFFFFF) % 10000
    return bucket > int(bad_ratio * 10000)


reports_engine = ReportsEngine()
